#!/usr/bin/env python3
import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "product": ["product"],
    "brief": ["brief", "campaignbrief"],
    "businessObjective": ["businessobjective", "objective"],
    "creativeObjectiveFunnel": ["creativeobjectivefunnel", "creativeobjective", "funnel", "creativefunnel"],
    "videoDuration": ["videoduration", "duration", "videolength"],
    "ratioDimensions": ["ratiodimensions", "ratio", "dimensions", "aspectratio", "ratioordimensions"],
    "language": ["language", "lang"],
    "notificationEmail": [
        "notificationemail",
        "recipientemail",
        "email",
        "emailaddress",
        "emailid",
        "notifyemail",
        "deliveryemail",
    ],
}


def compact(value):
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", compact(value).lower())


def read_rows(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    table = [[compact(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if not table:
        return []

    header_row_index = 0
    for row_index, row in enumerate(table[:10]):
        normalized = [normalize_header(cell) for cell in row]
        if any(item in HEADER_ALIASES["product"] for item in normalized) and any(
            item in HEADER_ALIASES["brief"] for item in normalized
        ):
            header_row_index = row_index
            break

    headers = [normalize_header(cell) for cell in table[header_row_index]]
    index_by_key = {}
    for key, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(headers):
            if header in aliases:
                index_by_key[key] = index
                break

    parsed = []
    for row_index, row in enumerate(table[header_row_index + 1 :], start=header_row_index + 2):
        if not any(compact(cell) for cell in row):
            continue

        def cell(key):
            index = index_by_key.get(key)
            if index is None or index >= len(row):
                return ""
            return compact(row[index])

        parsed.append(
            {
                "sheetRow": row_index,
                "product": cell("product"),
                "brief": cell("brief"),
                "businessObjective": cell("businessObjective"),
                "creativeObjectiveFunnel": cell("creativeObjectiveFunnel"),
                "videoDuration": cell("videoDuration"),
                "ratioDimensions": cell("ratioDimensions"),
                "language": cell("language"),
                "notificationEmail": cell("notificationEmail"),
            }
        )

    return parsed


def request_json(method, url, payload=None, timeout=300):
    data = None
    headers = {"Accept": "application/json"}
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            body = response.read().decode("utf-8")
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {error.code} from {url}: {body}") from error


def summarize_job(job):
    video_step = next((step for step in job.get("steps", []) if step.get("id") == "video_render"), {})
    renders = job.get("renders") or []
    render = renders[0] if renders else {}
    return {
        "jobStatus": job.get("status"),
        "stepStatus": video_step.get("status"),
        "message": video_step.get("message"),
        "renderStatus": render.get("status"),
        "requestId": render.get("requestId") or job.get("operationName"),
        "assetFile": render.get("assetFile"),
        "error": render.get("error") or video_step.get("error") or job.get("error"),
    }


def log(message):
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {message}", flush=True)


def main():
    parser = argparse.ArgumentParser(description="Run a Sora Studio XLSX batch sequentially.")
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument("--base-url", default="http://127.0.0.1:3000", help="Local app base URL")
    parser.add_argument("--poll-seconds", type=int, default=30, help="Seconds between status polls")
    parser.add_argument("--max-rows", type=int, default=300, help="Safety cap for usable rows")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    rows = read_rows(workbook_path)
    if len(rows) > args.max_rows:
        raise SystemExit(f"Workbook has {len(rows)} usable rows, above cap {args.max_rows}.")
    if not rows:
        raise SystemExit("No usable rows found.")

    jobs_url = urllib.parse.urljoin(args.base_url.rstrip("/") + "/", "api/sora-studio/jobs")
    completed = []
    failed = []

    log(f"BATCH_START file={workbook_path} total={len(rows)}")
    for item_number, row in enumerate(rows, start=1):
        recipient = row.get("notificationEmail") or "unassigned"
        product = row.get("product") or ""
        duration = row.get("videoDuration") or ""
        ratio = row.get("ratioDimensions") or ""
        language = row.get("language") or ""
        log(
            "ROW_START "
            f"row={row['sheetRow']} item={item_number}/{len(rows)} recipient={recipient} "
            f"product={product} duration={duration} ratio={ratio} language={language}"
        )

        payload = {
            "product": row.get("product", ""),
            "brief": row.get("brief", ""),
            "businessObjective": row.get("businessObjective", ""),
            "creativeObjectiveFunnel": row.get("creativeObjectiveFunnel", ""),
            "videoDuration": row.get("videoDuration", ""),
            "ratioDimensions": row.get("ratioDimensions", ""),
            "language": row.get("language", ""),
            "notificationEmail": row.get("notificationEmail", ""),
            "strictParityMode": True,
            "autoRender": True,
        }

        result = request_json("POST", jobs_url, payload)
        job = result["job"]
        job_id = job["id"]
        job_url = urllib.parse.urljoin(jobs_url.rstrip("/") + "/", job_id)
        log(f"ROW_JOB_CREATED row={row['sheetRow']} item={item_number}/{len(rows)} recipient={recipient} product={product} job={job_id}")

        while True:
            status_payload = request_json("GET", job_url)
            job = status_payload["job"]
            summary = summarize_job(job)
            log(
                "ROW_STATUS "
                f"row={row['sheetRow']} item={item_number}/{len(rows)} recipient={recipient} "
                f"product={product} job={job_id} {json.dumps(summary, ensure_ascii=False)}"
            )
            if job.get("status") == "completed":
                asset_file = summary.get("assetFile") or ""
                log(
                    "ROW_DONE "
                    f"row={row['sheetRow']} item={item_number}/{len(rows)} recipient={recipient} "
                    f"product={product} job={job_id} asset={asset_file}"
                )
                completed.append(
                    {
                        "sheetRow": row["sheetRow"],
                        "status": "completed",
                        "recipient": recipient,
                        "product": product,
                        "jobId": job_id,
                        "assetFile": asset_file,
                    }
                )
                break
            if job.get("status") == "failed":
                error = summary.get("error") or "Unknown error"
                log(
                    "ROW_FAILED "
                    f"row={row['sheetRow']} item={item_number}/{len(rows)} recipient={recipient} "
                    f"product={product} job={job_id} error={error}"
                )
                failed.append(
                    {
                        "sheetRow": row["sheetRow"],
                        "status": "failed",
                        "recipient": recipient,
                        "product": product,
                        "jobId": job_id,
                        "error": error,
                    }
                )
                break
            time.sleep(max(5, args.poll_seconds))

    log(f"BATCH_DONE total={len(rows)} completed={len(completed)} failed={len(failed)}")
    print(json.dumps(completed + failed, ensure_ascii=False, indent=2), flush=True)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
