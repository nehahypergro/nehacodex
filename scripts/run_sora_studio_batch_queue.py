#!/usr/bin/env python3
import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "product": ["product"],
    "brief": ["brief", "campaignbrief"],
    "businessObjective": ["businessobjective", "objective"],
    "creativeObjectiveFunnel": ["creativeobjectivefunnel", "creativeobjective", "funnel", "creativefunnel"],
    "videoDuration": ["videoduration", "duration", "videolength"],
    "ratioDimensions": ["ratiodimensions", "ratio", "dimensions", "aspectratio", "ratioordimensions"],
    "language": ["language", "lang"],
    "notificationEmail": [
        "notificationemail",
        "recipientemail",
        "email",
        "emailaddress",
        "emailid",
        "notifyemail",
        "deliveryemail",
    ],
}


def compact(value):
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", compact(value).lower())


def signature_for(product, brief, email, duration="", ratio="", language="", sheet_row=""):
    return (
        compact(product),
        compact(brief),
        compact(email),
        compact(duration),
        compact(ratio),
        compact(language),
        compact(sheet_row),
    )


def read_rows(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    table = [[compact(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if not table:
        return []

    header_row_index = 0
    for row_index, row in enumerate(table[:10]):
        normalized = [normalize_header(cell) for cell in row]
        if any(item in HEADER_ALIASES["product"] for item in normalized) and any(
            item in HEADER_ALIASES["brief"] for item in normalized
        ):
            header_row_index = row_index
            break

    headers = [normalize_header(cell) for cell in table[header_row_index]]
    index_by_key = {}
    for key, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(headers):
            if header in aliases:
                index_by_key[key] = index
                break

    parsed = []
    for row_index, row in enumerate(table[header_row_index + 1 :], start=header_row_index + 2):
        if not any(compact(cell) for cell in row):
            continue

        def cell(key):
            index = index_by_key.get(key)
            if index is None or index >= len(row):
                return ""
            return compact(row[index])

        item = {
            "sheetRow": row_index,
            "product": cell("product"),
            "brief": cell("brief"),
            "businessObjective": cell("businessObjective"),
            "creativeObjectiveFunnel": cell("creativeObjectiveFunnel"),
            "videoDuration": cell("videoDuration"),
            "ratioDimensions": cell("ratioDimensions"),
            "language": cell("language"),
            "notificationEmail": cell("notificationEmail"),
        }
        item["signature"] = signature_for(
            item["product"],
            item["brief"],
            item["notificationEmail"],
            item["videoDuration"],
            item["ratioDimensions"],
            item["language"],
        )
        item["rowSignature"] = signature_for(
            item["product"],
            item["brief"],
            item["notificationEmail"],
            item["videoDuration"],
            item["ratioDimensions"],
            item["language"],
            row_index,
        )
        parsed.append(item)

    return parsed


def request_json(method, url, payload=None, timeout=300):
    data = None
    headers = {"Accept": "application/json"}
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            body = response.read().decode("utf-8")
            return json.loads(body)
    except urllib.error.HTTPError as error:
        body = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {error.code} from {url}: {body}") from error


def summarize_job(job):
    video_step = next((step for step in job.get("steps", []) if step.get("id") == "video_render"), {})
    renders = job.get("renders") or []
    render = renders[0] if renders else {}
    return {
        "jobStatus": job.get("status"),
        "stepStatus": video_step.get("status"),
        "message": video_step.get("message"),
        "renderStatus": render.get("status"),
        "requestId": render.get("requestId") or job.get("operationName"),
        "assetFile": render.get("assetFile"),
        "error": render.get("error") or video_step.get("error") or job.get("error"),
    }


def is_terminal_failed_job(job):
    if job.get("status") == "failed":
        return True

    summary = summarize_job(job)
    error = compact(summary.get("error", ""))
    if not error or summary.get("assetFile"):
        return False

    video_step = next((step for step in job.get("steps", []) if step.get("id") == "video_render"), {})
    renders = job.get("renders") or []
    render = renders[0] if renders else {}
    lower_error = error.lower()
    return bool(
        video_step.get("completedAt")
        or render.get("completedAt")
        or "timed out" in lower_error
        or "all model renders failed" in lower_error
    )


def load_existing_jobs(root):
    jobs = []
    for job_file in root.glob("*/job.json"):
        try:
            job = json.loads(job_file.read_text())
        except Exception:
            continue
        item_input = job.get("input") or {}
        row_number = item_input.get("rowNumber", "")
        if row_number == 1:
            row_number = ""
        sig = signature_for(
            item_input.get("product", ""),
            item_input.get("brief", ""),
            item_input.get("notificationEmail", ""),
            item_input.get("videoDuration", ""),
            item_input.get("ratioDimensions", ""),
            item_input.get("language", ""),
        )
        row_sig = signature_for(
            item_input.get("product", ""),
            item_input.get("brief", ""),
            item_input.get("notificationEmail", ""),
            item_input.get("videoDuration", ""),
            item_input.get("ratioDimensions", ""),
            item_input.get("language", ""),
            row_number,
        )
        jobs.append(
            {
                "id": job.get("id", job_file.parent.name),
                "status": job.get("status", ""),
                "createdAt": job.get("createdAt", ""),
                "signature": sig,
                "rowSignature": row_sig,
                "job": job,
            }
        )
    jobs.sort(key=lambda item: item["createdAt"])
    return jobs


def existing_job_by_signature(root, min_created_at=""):
    selected = {}
    priority = {"completed": 3, "running": 2, "queued": 2, "failed": 1}
    for item in load_existing_jobs(root):
        if min_created_at and item["createdAt"] < min_created_at:
            continue
        for sig in {item["signature"], item["rowSignature"]}:
            current = selected.get(sig)
            if current is None:
                selected[sig] = item
                continue
            current_score = priority.get(current["status"], 0)
            item_score = priority.get(item["status"], 0)
            if item_score > current_score or (item_score == current_score and item["createdAt"] > current["createdAt"]):
                selected[sig] = item
    return selected


def log(message):
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {message}", flush=True)


def row_label(row, index, total):
    recipient = row.get("notificationEmail") or "unassigned"
    return (
        f"row={row['sheetRow']} item={index}/{total} recipient={recipient} "
        f"product={row.get('product', '')}"
    )


def create_job(row, jobs_url, attempts=3, retry_delay_seconds=30):
    payload = {
        "product": row.get("product", ""),
        "brief": row.get("brief", ""),
        "businessObjective": row.get("businessObjective", ""),
        "creativeObjectiveFunnel": row.get("creativeObjectiveFunnel", ""),
        "videoDuration": row.get("videoDuration", ""),
        "ratioDimensions": row.get("ratioDimensions", ""),
        "language": row.get("language", ""),
        "notificationEmail": row.get("notificationEmail", ""),
        "rowNumber": row.get("sheetRow"),
        "strictParityMode": True,
        "autoRender": True,
    }
    last_error = None
    for attempt in range(1, max(1, attempts) + 1):
        try:
            return request_json("POST", jobs_url, payload)["job"]
        except Exception as error:
            last_error = error
            if attempt >= max(1, attempts):
                break
            log(f"ROW_CREATE_RETRY {row_label(row, row['itemIndex'], row['totalItems'])} attempt={attempt} error={error}")
            time.sleep(max(1, retry_delay_seconds))
    raise RuntimeError(str(last_error or "Job creation failed."))


def record_create_failure(row, total, error, failed):
    failed.append(
        {
            "sheetRow": row["sheetRow"],
            "status": "create_failed",
            "recipient": row.get("notificationEmail") or "unassigned",
            "product": row.get("product", ""),
            "error": str(error),
        }
    )
    log(f"ROW_CREATE_FAILED {row_label(row, row['itemIndex'], total)} error={error}")


def main():
    parser = argparse.ArgumentParser(description="Run a Sora Studio XLSX batch with controlled concurrency.")
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument("--base-url", default="http://127.0.0.1:3000", help="Local app base URL")
    parser.add_argument("--concurrency", type=int, default=3, help="Number of rows to run at once")
    parser.add_argument("--poll-seconds", type=int, default=30, help="Seconds between status polls")
    parser.add_argument("--wait-existing-before-start", action="store_true", help="Wait for existing active jobs before launching new rows")
    parser.add_argument(
        "--ignore-existing-completed",
        action="store_true",
        help="Create jobs for every workbook row even when an older completed job has the same product, brief, and recipient.",
    )
    parser.add_argument(
        "--current-batch-start-iso",
        default="",
        help="Only adopt/skip existing jobs created at or after this ISO timestamp.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    rows = read_rows(workbook_path)
    if not rows:
        raise SystemExit("No usable rows found.")

    jobs_url = urllib.parse.urljoin(args.base_url.rstrip("/") + "/", "api/sora-studio/jobs")
    generated_root = Path.cwd() / "generated-sora-studio"
    existing_by_sig = existing_job_by_signature(generated_root, args.current_batch_start_iso.strip())

    active = []
    pending = []
    completed = []
    failed = []

    total = len(rows)
    for index, row in enumerate(rows, start=1):
        row["itemIndex"] = index
        row["totalItems"] = total
        existing = existing_by_sig.get(row.get("rowSignature")) or existing_by_sig.get(row["signature"])
        if existing and is_terminal_failed_job(existing["job"]):
            summary = summarize_job(existing["job"])
            error = summary.get("error") or existing["job"].get("error") or "Unknown error"
            failed.append(
                {
                    "sheetRow": row["sheetRow"],
                    "status": "failed_existing",
                    "recipient": row.get("notificationEmail") or "unassigned",
                    "product": row.get("product", ""),
                    "jobId": existing["id"],
                    "error": error,
                }
            )
            log(f"ROW_FAILED_EXISTING {row_label(row, index, total)} job={existing['id']} error={error}")
        elif existing and existing["status"] == "completed" and not args.ignore_existing_completed:
            summary = summarize_job(existing["job"])
            completed.append(
                {
                    "sheetRow": row["sheetRow"],
                    "status": "completed_existing",
                    "recipient": row.get("notificationEmail") or "unassigned",
                    "product": row.get("product", ""),
                    "jobId": existing["id"],
                    "assetFile": summary.get("assetFile") or "",
                }
            )
            log(f"ROW_SKIP_COMPLETED {row_label(row, index, total)} job={existing['id']} asset={summary.get('assetFile') or ''}")
        elif existing and existing["status"] in {"running", "queued"}:
            active.append({"row": row, "jobId": existing["id"], "existing": True})
            log(f"ROW_ADOPT_ACTIVE {row_label(row, index, total)} job={existing['id']}")
        else:
            pending.append(row)

    log(
        "BATCH_QUEUE_START "
        f"file={workbook_path} total={total} completed_existing={len(completed)} "
        f"active_existing={len(active)} pending_new={len(pending)} concurrency={args.concurrency}"
    )

    if args.wait_existing_before_start and active:
        log(f"WAIT_EXISTING_ACTIVE count={len(active)} before_starting_new=true")

    def maybe_start_jobs():
        if args.wait_existing_before_start and any(item.get("existing") for item in active):
            return
        while pending and len(active) < max(1, args.concurrency):
            row = pending.pop(0)
            index = row["itemIndex"]
            log(
                "ROW_START "
                f"{row_label(row, index, total)} duration={row.get('videoDuration', '')} "
                f"ratio={row.get('ratioDimensions', '')} language={row.get('language', '')}"
            )
            try:
                job = create_job(row, jobs_url)
            except Exception as error:
                record_create_failure(row, total, error, failed)
                continue
            job_id = job["id"]
            log(f"ROW_JOB_CREATED {row_label(row, index, total)} job={job_id}")
            active.append({"row": row, "jobId": job_id, "existing": False})

    maybe_start_jobs()

    while active or pending:
        next_active = []
        for item in active:
            row = item["row"]
            index = row["itemIndex"]
            job_id = item["jobId"]
            job_url = urllib.parse.urljoin(jobs_url.rstrip("/") + "/", job_id)
            status_payload = request_json("GET", job_url)
            job = status_payload["job"]
            summary = summarize_job(job)
            log(f"ROW_STATUS {row_label(row, index, total)} job={job_id} {json.dumps(summary, ensure_ascii=False)}")
            if job.get("status") == "completed":
                asset_file = summary.get("assetFile") or ""
                log(f"ROW_DONE {row_label(row, index, total)} job={job_id} asset={asset_file}")
                completed.append(
                    {
                        "sheetRow": row["sheetRow"],
                        "status": "completed",
                        "recipient": row.get("notificationEmail") or "unassigned",
                        "product": row.get("product", ""),
                        "jobId": job_id,
                        "assetFile": asset_file,
                    }
                )
            elif job.get("status") == "failed" or is_terminal_failed_job(job):
                error = summary.get("error") or "Unknown error"
                log(f"ROW_FAILED {row_label(row, index, total)} job={job_id} error={error}")
                failed.append(
                    {
                        "sheetRow": row["sheetRow"],
                        "status": "failed",
                        "recipient": row.get("notificationEmail") or "unassigned",
                        "product": row.get("product", ""),
                        "jobId": job_id,
                        "error": error,
                    }
                )
            else:
                next_active.append(item)

        active = next_active
        maybe_start_jobs()
        if active or pending:
            time.sleep(max(5, args.poll_seconds))

    log(f"BATCH_QUEUE_DONE total={total} completed_or_existing={len(completed)} failed={len(failed)}")
    print(json.dumps(completed + failed, ensure_ascii=False, indent=2), flush=True)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
